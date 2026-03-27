"""
excel_report.py — McKinsey-style Excel presentation workbook
=============================================================
Generates a publication/presentation-ready .xlsx file with:
  Sheet 1: Overview      — executive summary + key findings
  Sheet 2: Patient Flow  — attrition table
  Sheet 3: Baseline      — Table 1 characteristics
  Sheet 4: OS Analysis   — KM curve + supporting table + number at risk
  Sheet 5: Methodology   — full methods text
"""

import datetime
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from config import (
    OUTPUT_DIR,
    DIAGNOSIS_START, DIAGNOSIS_END, FOLLOWUP_CUTOFF,
    LANDMARK_MONTHS,
    FIXED_DURATION_LOWER_MONTHS, FIXED_DURATION_UPPER_MONTHS,
    CONTINUATION_LOWER_MONTHS,
    MAX_INFUSION_GAP_DAYS, CHEMO_PROXIMITY_EXCLUSION_DAYS,
    COVARIATES_ADJUSTED,
)

# ── Colour palette ────────────────────────────────────────────────────────────
C_NAVY        = "1F3864"
C_BLUE        = "2E75B6"
C_MID_BLUE    = "4472C4"
C_LIGHT_BLUE  = "D6E4F7"
C_FD          = "2166AC"   # Fixed-Duration cohort
C_CONT        = "C00000"   # Continuation cohort
C_WHITE       = "FFFFFF"
C_OFF_WHITE   = "F7F9FC"
C_LIGHT_GRAY  = "F2F2F2"
C_MID_GRAY    = "D9D9D9"
C_DARK_GRAY   = "595959"
C_BORDER      = "BFBFBF"
C_GOLD        = "C9A227"


# ── Style helpers ─────────────────────────────────────────────────────────────

def _font(name="Calibri", size=10, bold=False, italic=False,
          color="000000", underline=None):
    return Font(name=name, size=size, bold=bold, italic=italic,
                color=color, underline=underline)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _thin_border(all_sides=True, bottom_only=False, top_bottom=False):
    thin = Side(style="thin", color=C_BORDER)
    none = Side(style=None)
    if bottom_only:
        return Border(bottom=thin)
    if top_bottom:
        return Border(top=thin, bottom=thin)
    if all_sides:
        return Border(left=thin, right=thin, top=thin, bottom=thin)
    return Border()


def _align(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


def _col(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def _row_h(ws, row_idx, height):
    ws.row_dimensions[row_idx].height = height


def _cell(ws, r, c, value="", font=None, fill=None,
          align=None, border=None, num_fmt=None):
    cell = ws.cell(row=r, column=c, value=value)
    if font:    cell.font = font
    if fill:    cell.fill = fill
    if align:   cell.alignment = align
    if border:  cell.border = border
    if num_fmt: cell.number_format = num_fmt
    return cell


def _merge(ws, r, c1, c2, value="", font=None, fill=None,
           align=None, border=None):
    ws.merge_cells(start_row=r, start_column=c1,
                   end_row=r, end_column=c2)
    return _cell(ws, r, c1, value, font, fill, align, border)


def _apply_all_borders(ws, r1, c1, r2, c2):
    """Apply thin borders to every cell in a rectangle."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = _thin_border()


def _no_gridlines(ws):
    ws.sheet_view.showGridLines = False


def _format_adjusted_covariates(covariates):
    """Convert adjusted model covariate names into presentation-friendly labels."""
    label_map = {
        "age_at_index": "age",
        "race": "race",
        "ecog_binary": "ECOG",
        "pdl1_cat": "PD-L1",
        "histology_cat": "histology",
        "pembro_with_chemo": "treatment type",
    }
    labels = [label_map.get(cov, cov.replace("_", " ")) for cov in covariates]
    return ", ".join(labels)


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_overview(wb, cohort_df, attrition, km_output):
    ws = wb.create_sheet("Overview")
    ws.sheet_properties.tabColor = C_NAVY
    _no_gridlines(ws)

    # Column widths
    widths = [2, 28, 22, 22, 2]
    for i, w in enumerate(widths, 1):
        _col(ws, i, w)

    # ── Banner ────────────────────────────────────────────────────────────
    _row_h(ws, 1, 8)
    _merge(ws, 2, 1, 5, "",
           fill=_fill(C_NAVY))

    _row_h(ws, 3, 36)
    _merge(ws, 3, 2, 4,
           "Overall Survival After Fixed-Duration vs Continuation\nof Pembrolizumab in First-Line Metastatic NSCLC",
           font=_font(size=16, bold=True, color=C_WHITE),
           fill=_fill(C_NAVY),
           align=_align("left", "center", wrap=True))

    _row_h(ws, 4, 20)
    _merge(ws, 4, 2, 4,
           "A Landmark Analysis Using IC PrecisionQ Real-World EHR Data",
           font=_font(size=11, italic=True, color=C_LIGHT_BLUE),
           fill=_fill(C_NAVY),
           align=_align("left", "center"))

    _row_h(ws, 5, 8)
    _merge(ws, 5, 1, 5, "", fill=_fill(C_NAVY))

    _row_h(ws, 6, 10)

    # ── Key metrics row ───────────────────────────────────────────────────
    n_fd   = int((cohort_df["cohort"] == "Fixed-Duration").sum())
    n_cont = int((cohort_df["cohort"] == "Continuation").sum())
    n_tot  = len(cohort_df)
    lr     = km_output["logrank"]
    p_val  = f"{lr.p_value:.4f}" if lr is not None else "N/A"
    summary = km_output["summary"]

    def _get_median(cohort_name):
        row = summary[summary["Cohort"] == cohort_name]
        if row.empty:
            return "NR"
        m = row.iloc[0]["Median OS (months)"]
        lo = row.iloc[0]["95% CI Lower"]
        hi = row.iloc[0]["95% CI Upper"]
        if pd.isna(m):
            return "NR"
        return f"{m:.1f} mo\n(95% CI: {lo:.1f}-{hi:.1f})"

    metrics = [
        ("Analysis Cohort", f"{n_tot:,} patients", C_NAVY),
        ("Fixed-Duration", f"N = {n_fd:,}\nMedian OS: {_get_median('Fixed-Duration')}", C_FD),
        ("Continuation", f"N = {n_cont:,}\nMedian OS: {_get_median('Continuation')}", C_CONT),
        ("Log-rank P-value", p_val, C_DARK_GRAY),
    ]

    # Metric boxes: col B (2) to D (4), two boxes per row
    box_positions = [(7, 2), (7, 4), (10, 2), (10, 4)]
    for idx, ((r, c), (label, value, color)) in enumerate(zip(box_positions, metrics)):
        _row_h(ws, r, 18)
        _row_h(ws, r + 1, 28)
        _row_h(ws, r + 2, 6)

        _merge(ws, r, c, c, label,
               font=_font(size=9, bold=True, color=C_WHITE),
               fill=_fill(color),
               align=_align("center", "center"))

        _merge(ws, r + 1, c, c, value,
               font=_font(size=11, bold=True, color=color),
               fill=_fill(C_OFF_WHITE),
               align=_align("center", "center", wrap=True))

        # Box border
        for rr in [r, r + 1]:
            for cc in [c]:
                ws.cell(row=rr, column=cc).border = _thin_border()

    _row_h(ws, 13, 12)

    # ── Study parameters section ──────────────────────────────────────────
    _row_h(ws, 14, 18)
    _merge(ws, 14, 2, 4, "STUDY PARAMETERS",
           font=_font(size=10, bold=True, color=C_WHITE),
           fill=_fill(C_BLUE),
           align=_align("left", "center", indent=1))

    params = [
        ("Data Source",        "IC PrecisionQ NSCLC — Real-World EHR"),
        ("Study Design",       "Retrospective observational cohort, landmark analysis"),
        ("Diagnosis Window",   f"{DIAGNOSIS_START} to {DIAGNOSIS_END}"),
        ("Follow-up Cutoff",   str(FOLLOWUP_CUTOFF)),
        ("Landmark Timepoint", f"{LANDMARK_MONTHS} months from pembrolizumab initiation"),
        ("Fixed-Duration",     f"Last infusion {FIXED_DURATION_LOWER_MONTHS}-{FIXED_DURATION_UPPER_MONTHS} months from start"),
        ("Continuation",       f"Last infusion >{CONTINUATION_LOWER_MONTHS} months from start"),
        ("Infusion Gap Rule",  f"Treatment ended if >={MAX_INFUSION_GAP_DAYS} days between infusions"),
    ]

    for i, (param, val) in enumerate(params):
        r = 15 + i
        _row_h(ws, r, 16)
        fill = _fill(C_OFF_WHITE) if i % 2 == 0 else _fill(C_WHITE)
        _cell(ws, r, 2, param,
              font=_font(size=10, bold=True, color=C_DARK_GRAY),
              fill=fill,
              align=_align("left", "center", indent=1),
              border=_thin_border(bottom_only=True))
        _cell(ws, r, 3, val,
              font=_font(size=10),
              fill=fill,
              align=_align("left", "center"),
              border=_thin_border(bottom_only=True))
        ws.merge_cells(start_row=r, start_column=3,
                       end_row=r, end_column=4)

    # ── Footer ────────────────────────────────────────────────────────────
    footer_r = 15 + len(params) + 2
    _row_h(ws, footer_r, 14)
    _merge(ws, footer_r, 2, 4,
           f"CONFIDENTIAL  |  Analysis date: {datetime.date.today().isoformat()}  |  "
           f"For research purposes only",
           font=_font(size=8, italic=True, color=C_DARK_GRAY),
           align=_align("center", "center"))


def _build_patient_flow(wb, attrition):
    ws = wb.create_sheet("Patient Flow")
    ws.sheet_properties.tabColor = C_DARK_GRAY
    _no_gridlines(ws)

    _col(ws, 1, 3)
    _col(ws, 2, 52)
    _col(ws, 3, 18)
    _col(ws, 4, 18)
    _col(ws, 5, 3)

    # Title
    _row_h(ws, 1, 8)
    _row_h(ws, 2, 30)
    _merge(ws, 2, 2, 4, "Patient Attrition Flow",
           font=_font(size=14, bold=True, color=C_WHITE),
           fill=_fill(C_NAVY),
           align=_align("left", "center", indent=1))
    _row_h(ws, 3, 16)
    _merge(ws, 3, 2, 4,
           "Sequential application of SAP inclusion/exclusion criteria",
           font=_font(size=10, italic=True, color=C_DARK_GRAY),
           align=_align("left", "center", indent=1))
    _row_h(ws, 4, 10)

    # Header row
    _row_h(ws, 5, 20)
    for c, (hdr, aln) in enumerate([
        ("Criteria / Step", "left"),
        ("Patients, N", "right"),
        ("Excluded, N (%)", "right"),
    ], 2):
        _cell(ws, 5, c, hdr,
              font=_font(size=10, bold=True, color=C_WHITE),
              fill=_fill(C_BLUE),
              align=_align(aln, "center", indent=1 if aln == "left" else 0),
              border=_thin_border())

    steps = [
        ("total_patients",          "Total patients in dataset",                 None),
        ("age_18_plus",             "Age >=18 years",                            None),
        ("metastatic_nsclc_c34x",   "Metastatic NSCLC (ICD-10: C34.X)",          None),
        ("diagnosis_window",        f"Diagnosed {DIAGNOSIS_START} to {DIAGNOSIS_END}", None),
        ("pembro_1l_metastatic",    "Pembrolizumab in 1st-line metastatic LOT",  None),
        ("has_infusion_data",       "With infusion-level dose data",             None),
        ("infusion_at_22_26_months", f"Infusion documented at >={FIXED_DURATION_LOWER_MONTHS} months",  None),
        ("after_chemo_exclusion",   f"After chemo proximity exclusion (>{CHEMO_PROXIMITY_EXCLUSION_DAYS}d of last pembro)", None),
        ("alive_at_landmark",       f"Alive at {LANDMARK_MONTHS}-month landmark",    None),
    ]

    keys = [s[0] for s in steps]
    prev_n = None
    for i, (key, label, _) in enumerate(steps):
        r = 6 + i
        _row_h(ws, r, 18)
        n = attrition.get(key)
        fill = _fill(C_OFF_WHITE) if i % 2 == 0 else _fill(C_WHITE)

        # Highlight final row
        if key == "alive_at_landmark":
            fill = _fill(C_LIGHT_BLUE)

        _cell(ws, r, 2, label,
              font=_font(size=10, bold=(key == "alive_at_landmark")),
              fill=fill,
              align=_align("left", "center", indent=1),
              border=_thin_border())

        n_str = f"{n:,}" if n is not None else "—"
        _cell(ws, r, 3, n_str,
              font=_font(size=10, bold=(key == "alive_at_landmark")),
              fill=fill,
              align=_align("right", "center"),
              border=_thin_border())

        if prev_n is not None and n is not None and prev_n > 0:
            excl = prev_n - n
            pct  = excl / prev_n * 100
            excl_str = f"{excl:,} ({pct:.1f}%)" if excl > 0 else "—"
        else:
            excl_str = "—"

        _cell(ws, r, 4, excl_str,
              font=_font(size=10, color=(C_CONT if excl_str != "—" else "000000")),
              fill=fill,
              align=_align("right", "center"),
              border=_thin_border())

        if n is not None:
            prev_n = n

    # Cohort split
    fd   = attrition.get("fixed_duration", 0)
    cont = attrition.get("continuation", 0)
    tot_cohort = (fd or 0) + (cont or 0)

    r = 6 + len(steps) + 1
    _row_h(ws, r, 8)
    r += 1

    for cohort_name, n_cohort, color in [
        ("Fixed-Duration Cohort", fd, C_FD),
        ("Continuation Cohort",   cont, C_CONT),
    ]:
        _row_h(ws, r, 20)
        pct = n_cohort / tot_cohort * 100 if tot_cohort > 0 else 0
        _cell(ws, r, 2, f"  {cohort_name}",
              font=_font(size=10, bold=True, color=C_WHITE),
              fill=_fill(color),
              align=_align("left", "center"),
              border=_thin_border())
        _cell(ws, r, 3, f"{n_cohort:,} ({pct:.1f}%)",
              font=_font(size=10, bold=True, color=C_WHITE),
              fill=_fill(color),
              align=_align("right", "center"),
              border=_thin_border())
        _cell(ws, r, 4, "",
              fill=_fill(color),
              border=_thin_border())
        r += 1


def _build_baseline(wb, table1_df):
    ws = wb.create_sheet("Baseline Characteristics")
    ws.sheet_properties.tabColor = C_MID_BLUE
    _no_gridlines(ws)

    _col(ws, 1, 3)
    _col(ws, 2, 38)

    # Detect group columns (all except "Variable")
    group_cols = [c for c in table1_df.columns if c != "Variable"]
    col_map = {grp: 3 + i for i, grp in enumerate(group_cols)}
    for grp, c in col_map.items():
        _col(ws, c, 28)
    _col(ws, 3 + len(group_cols), 3)

    # Title
    _row_h(ws, 1, 8)
    _row_h(ws, 2, 30)
    last_c = 2 + len(group_cols)
    _merge(ws, 2, 2, last_c, "Table 1: Baseline Characteristics at Landmark",
           font=_font(size=14, bold=True, color=C_WHITE),
           fill=_fill(C_NAVY),
           align=_align("left", "center", indent=1))
    _row_h(ws, 3, 16)
    _merge(ws, 3, 2, last_c,
           "Evaluated at the 29-month landmark date. N (%) for categorical; Median [IQR] for continuous.",
           font=_font(size=10, italic=True, color=C_DARK_GRAY),
           align=_align("left", "center", indent=1))
    _row_h(ws, 4, 10)

    # Column headers
    _row_h(ws, 5, 22)
    _cell(ws, 5, 2, "Characteristic",
          font=_font(size=10, bold=True, color=C_WHITE),
          fill=_fill(C_NAVY),
          align=_align("left", "center", indent=1),
          border=_thin_border())

    cohort_colors = {"Fixed-Duration": C_FD, "Continuation": C_CONT}
    for grp, c in col_map.items():
        color = cohort_colors.get(grp, C_BLUE)
        _cell(ws, 5, c, grp,
              font=_font(size=10, bold=True, color=C_WHITE),
              fill=_fill(color),
              align=_align("center", "center"),
              border=_thin_border())

    # Data rows
    for i, (_, row_data) in enumerate(table1_df.iterrows()):
        r = 6 + i
        _row_h(ws, r, 16)
        var_val = row_data["Variable"]

        is_header = (
            isinstance(var_val, str) and
            not var_val.startswith("  ") and
            all(
                str(v) == "" or (isinstance(v, float) and np.isnan(v))
                for v in row_data[group_cols]
            )
        )
        is_total = isinstance(var_val, str) and var_val.strip() == "N"

        if is_total:
            fill = _fill(C_LIGHT_BLUE)
            font_var = _font(size=10, bold=True)
            font_val = _font(size=10, bold=True)
        elif is_header:
            fill = _fill(C_LIGHT_GRAY)
            font_var = _font(size=10, bold=True, color=C_DARK_GRAY)
            font_val = _font(size=10, bold=True, color=C_DARK_GRAY)
        else:
            fill = _fill(C_OFF_WHITE) if i % 2 == 0 else _fill(C_WHITE)
            font_var = _font(size=10)
            font_val = _font(size=10)

        _cell(ws, r, 2, var_val,
              font=font_var, fill=fill,
              align=_align("left", "center", indent=1),
              border=_thin_border(bottom_only=True))

        for grp, c in col_map.items():
            val = row_data.get(grp, "")
            val_str = "" if (val is None or (isinstance(val, float) and np.isnan(val))) else str(val)
            _cell(ws, r, c, val_str,
                  font=font_val, fill=fill,
                  align=_align("center", "center"),
                  border=_thin_border(bottom_only=True))

    # Freeze header
    ws.freeze_panes = "B6"


def _build_os_analysis(wb, cohort_df, km_output, km_supporting):
    ws = wb.create_sheet("Overall Survival")
    ws.sheet_properties.tabColor = C_FD
    _no_gridlines(ws)

    _col(ws, 1, 3)
    _col(ws, 2, 32)
    _col(ws, 3, 26)
    _col(ws, 4, 26)
    _col(ws, 5, 3)

    # Title
    _row_h(ws, 1, 8)
    _row_h(ws, 2, 30)
    _merge(ws, 2, 2, 4, "Overall Survival from 29-Month Landmark",
           font=_font(size=14, bold=True, color=C_WHITE),
           fill=_fill(C_NAVY),
           align=_align("left", "center", indent=1))
    _row_h(ws, 3, 16)
    _merge(ws, 3, 2, 4,
           "Fixed-Duration vs Continuation Pembrolizumab — Kaplan-Meier Analysis",
           font=_font(size=10, italic=True, color=C_DARK_GRAY),
           align=_align("left", "center", indent=1))
    _row_h(ws, 4, 8)

    # ── Embed KM image ────────────────────────────────────────────────────
    km_img_path = Path(OUTPUT_DIR) / "km_os_landmark.png"
    img_start_row = 5
    img_rows = 22  # image takes ~22 rows

    if km_img_path.exists():
        try:
            img = XLImage(str(km_img_path))
            img.width  = 560
            img.height = 390
            ws.add_image(img, f"B{img_start_row}")
            for r in range(img_start_row, img_start_row + img_rows):
                _row_h(ws, r, 18)
        except Exception:
            _merge(ws, img_start_row, 2, 4, "[KM chart — km_os_landmark.png]",
                   font=_font(size=10, italic=True, color=C_DARK_GRAY),
                   align=_align("center", "center"))
    else:
        _merge(ws, img_start_row, 2, 4, "[KM chart not found — run analysis first]",
               font=_font(size=10, italic=True, color=C_DARK_GRAY),
               align=_align("center", "center"))

    sep_row = img_start_row + img_rows + 1

    # ── Supporting table ──────────────────────────────────────────────────
    _row_h(ws, sep_row, 18)
    _merge(ws, sep_row, 2, 4, "SURVIVAL SUMMARY",
           font=_font(size=10, bold=True, color=C_WHITE),
           fill=_fill(C_BLUE),
           align=_align("left", "center", indent=1))

    st = km_supporting["summary_table"]
    # Separate data rows from the log-rank footer row
    data_rows = st[st["Cohort"].notna() & ~st["Cohort"].astype(str).str.startswith("Log-rank")]
    footer_rows = st[st["Cohort"].astype(str).str.startswith("Log-rank")]

    # ── Transpose: metrics as rows, cohorts as columns ────────────────────
    cohort_names = list(data_rows["Cohort"].values)
    st_T = data_rows.set_index("Cohort").T.reset_index()
    st_T.columns = ["Metric"] + cohort_names

    cohort_colors_st = {"Fixed-Duration": C_FD, "Continuation": C_CONT}

    # Column widths for transposed table
    _col(ws, 2, 36)   # Metric column
    for ci, grp in enumerate(cohort_names):
        _col(ws, 3 + ci, 28)

    header_r = sep_row + 1
    _row_h(ws, header_r, 20)
    ncols_T = len(st_T.columns)  # "Metric" + cohorts

    # Header row: "Metric" + cohort names
    _cell(ws, header_r, 2, "Metric",
          font=_font(size=9, bold=True, color=C_WHITE),
          fill=_fill(C_NAVY),
          align=_align("left", "center", wrap=True, indent=1),
          border=_thin_border())
    for ci, grp in enumerate(cohort_names):
        color = cohort_colors_st.get(str(grp), C_NAVY)
        _cell(ws, header_r, 3 + ci, str(grp),
              font=_font(size=9, bold=True, color=C_WHITE),
              fill=_fill(color),
              align=_align("center", "center"),
              border=_thin_border())

    # Data rows (one per metric)
    for ri, (_, row_data) in enumerate(st_T.iterrows()):
        r = header_r + 1 + ri
        _row_h(ws, r, 18)
        fill = _fill(C_OFF_WHITE) if ri % 2 == 0 else _fill(C_WHITE)
        metric_name = str(row_data["Metric"])
        _cell(ws, r, 2, metric_name,
              font=_font(size=9, bold=True, color=C_DARK_GRAY),
              fill=fill,
              align=_align("left", "center", indent=1),
              border=_thin_border(bottom_only=True))
        for ci, grp in enumerate(cohort_names):
            val = row_data.get(grp, "")
            val_str = "" if (val is None or (isinstance(val, float) and np.isnan(val))) else str(val)
            cohort_color = cohort_colors_st.get(str(grp), "000000")
            _cell(ws, r, 3 + ci, val_str,
                  font=_font(size=9, color=cohort_color),
                  fill=fill,
                  align=_align("center", "center"),
                  border=_thin_border(bottom_only=True))

    # Log-rank p-value footer
    if not footer_rows.empty:
        r = header_r + 1 + len(st_T)
        _row_h(ws, r, 16)
        p_text = str(footer_rows.iloc[0]["Cohort"])
        _merge(ws, r, 2, 2 + ncols_T - 1, p_text,
               font=_font(size=9, bold=True, italic=True),
               fill=_fill(C_LIGHT_GRAY),
               align=_align("center", "center"),
               border=_thin_border(top_bottom=True))


def _build_methodology(wb, cohort_df, attrition, km_output):
    ws = wb.create_sheet("Methodology")
    ws.sheet_properties.tabColor = C_MID_GRAY
    _no_gridlines(ws)

    _col(ws, 1, 3)
    _col(ws, 2, 90)
    _col(ws, 3, 3)

    lr    = km_output["logrank"]
    p_val = f"{lr.p_value:.4f}" if lr is not None else "N/A"

    sections = [
        ("METHODOLOGY SUMMARY", None, "title"),
        ("Study Design", None, "section"),
        (None,
         "Retrospective observational cohort study using IC PrecisionQ de-identified oncology EHR data. "
         "A landmark analysis design was implemented to mitigate immortal time bias introduced by "
         "defining exposure groups based on treatment duration.",
         "body"),
        ("Data Source", None, "section"),
        (None,
         f"IC PrecisionQ NSCLC dataset. Follow-up through {FOLLOWUP_CUTOFF}. "
         f"Diagnosis window: {DIAGNOSIS_START} to {DIAGNOSIS_END}.",
         "body"),
        ("Study Population — Inclusion Criteria", None, "section"),
        (None, "  - Age >=18 years at diagnosis", "body"),
        (None, "  - Metastatic NSCLC (ICD-10: C34.X), de novo or recurrent", "body"),
        (None, f"  - Diagnosis date between {DIAGNOSIS_START} and {DIAGNOSIS_END}", "body"),
        (None, "  - Pembrolizumab in first metastatic line of therapy (with or without chemotherapy)", "body"),
        (None, f"  - At least one pembrolizumab infusion at >={FIXED_DURATION_LOWER_MONTHS} months from treatment start", "body"),
        (None, f"  - Alive at the {LANDMARK_MONTHS}-month landmark", "body"),
        ("Study Population — Exclusion Criteria", None, "section"),
        (None,
         f"  - Chemotherapy administered within {CHEMO_PROXIMITY_EXCLUSION_DAYS} days of the last pembrolizumab infusion "
         "(interpreted as a progression-motivated treatment stop, excluded to reduce confounding).",
         "body"),
        ("Treatment Gap Rule", None, "section"),
        (None,
         f"If two consecutive pembrolizumab infusions were separated by more than {MAX_INFUSION_GAP_DAYS} days (~6 months), "
         "pembrolizumab treatment was considered to have ended at the earlier infusion date. "
         "This rule prevents gap-and-restart patterns from artificially extending treatment duration.",
         "body"),
        ("Exposure Groups", None, "section"),
        (None,
         f"  Fixed-Duration: Patients whose last pembrolizumab infusion (by gap rule) occurred between "
         f"{FIXED_DURATION_LOWER_MONTHS} and {FIXED_DURATION_UPPER_MONTHS} months from treatment initiation.",
         "body"),
        (None,
         f"  Continuation: Patients who continued pembrolizumab for more than {CONTINUATION_LOWER_MONTHS} months from initiation.",
         "body"),
        ("Landmark Analysis", None, "section"),
        (None,
         f"The landmark time is defined as {LANDMARK_MONTHS} months following pembrolizumab initiation. "
         "Only patients confirmed alive and under active observation at the landmark are included. "
         "Survival time is measured from the landmark date, not from treatment initiation.",
         "body"),
        ("Primary Endpoint", None, "section"),
        (None,
         "Overall survival (OS): time from the landmark date to death from any cause. "
         f"Patients alive at the end of follow-up are censored at date of last clinical contact "
         f"(administrative censoring at {FOLLOWUP_CUTOFF}).",
         "body"),
        ("Statistical Methods", None, "section"),
        (None, "  - Kaplan-Meier method for survival function estimation", "body"),
        (None, "  - Log-rank test for between-group comparison (two-sided, alpha = 0.05)", "body"),
        (None, f"  - Observed log-rank p-value: {p_val}", "body"),
        (None, "  - Median follow-up estimated using reverse Kaplan-Meier (Schemper & Smith method)", "body"),
        (None, "  - Multivariable Cox proportional hazards (CoxPH) models with sequential confounder adjustment", "body"),
        (None, "  - Schoenfeld residuals test for proportional hazards (PH) assumption (global + per covariate)", "body"),
        (None, "  - Subgroup analyses within 7 pre-specified subgroups with treatment-by-subgroup interaction testing", "body"),
        (None, "  - Landmark sensitivity: treatment HR re-estimated at 27, 29, and 32-month landmarks", "body"),
        (None, "  - LASSO-penalized Cox (L1, cross-validated alpha) for data-driven covariate selection", "body"),
        ("Cox Model Hierarchy", None, "section"),
        (None, "  Model 1 (Unadjusted): Treatment indicator only", "body"),
        (None, "  Model 2 (Minimal): + Age at landmark, sex", "body"),
        (None, "  Model 3 (Clinical): + ECOG, PD-L1, histology, brain mets, de novo vs recurrent, pembro ± chemo, smoking", "body"),
        (None, "  Model 4 (Full): Model 3 + comorbidities (diabetes, respiratory, cardiac, kidney) + medication proxies", "body"),
        (None, "  Model 5 (LASSO): LASSO-selected subset of Model 4 covariates (data-driven)", "body"),
        ("Software", None, "section"),
        (None, "Python 3.x with lifelines, pandas, matplotlib, openpyxl", "body"),
        ("References", None, "section"),
        (None, "1. Rousseau A et al. Lancet Reg Health Eur 2024;43:100970", "body"),
        (None, "2. Sasaki T et al. J Rural Med 2024;19(4):273-278", "body"),
        (None, "3. Sun L et al. JAMA Oncol 2023;9(8):1075-1082", "body"),
    ]

    r = 2
    _row_h(ws, 1, 8)

    for label, text, style in sections:
        if style == "title":
            _row_h(ws, r, 30)
            _merge(ws, r, 2, 2, label,
                   font=_font(size=14, bold=True, color=C_WHITE),
                   fill=_fill(C_NAVY),
                   align=_align("left", "center", indent=1))
            r += 1
            _row_h(ws, r, 6)
            r += 1

        elif style == "section":
            _row_h(ws, r, 8)
            r += 1
            _row_h(ws, r, 20)
            _cell(ws, r, 2, label,
                  font=_font(size=11, bold=True, color=C_WHITE),
                  fill=_fill(C_BLUE),
                  align=_align("left", "center", indent=1),
                  border=_thin_border(bottom_only=True))
            r += 1

        elif style == "body":
            _row_h(ws, r, 16)
            content = text or ""
            _cell(ws, r, 2, content,
                  font=_font(size=10),
                  align=_align("left", "center", wrap=True, indent=1),
                  border=_thin_border(bottom_only=True))
            r += 1

    # Footer
    r += 1
    _row_h(ws, r, 16)
    _cell(ws, r, 2,
          f"Report generated: {datetime.date.today().isoformat()}  |  For research purposes only",
          font=_font(size=8, italic=True, color=C_DARK_GRAY),
          align=_align("left", "center", indent=1))


# ── Main entry point ──────────────────────────────────────────────────────────

def _build_cox_comparison(wb, comparison_df, cox_results=None):
    """Sheet: Cox Model Comparison — treatment HR across model specifications."""
    ws = wb.create_sheet("Cox Model Comparison")
    ws.sheet_properties.tabColor = "B2182B"
    _no_gridlines(ws)

    _col(ws, 1, 3)
    _col(ws, 2, 22)
    for c in range(3, 10):
        _col(ws, c, 18)
    _col(ws, 10, 3)

    # Title
    _row_h(ws, 1, 8)
    _row_h(ws, 2, 30)
    _merge(ws, 2, 2, 9, "Cox Model Comparison — Treatment HR (Continuation vs Fixed-Duration)",
           font=_font(size=14, bold=True, color=C_WHITE),
           fill=_fill(C_NAVY),
           align=_align("left", "center", indent=1))
    _row_h(ws, 3, 16)
    _merge(ws, 3, 2, 9,
           "Sensitivity to confounder adjustment. HR <1 favours Continuation; HR >1 favours Fixed-Duration.",
           font=_font(size=10, italic=True, color=C_DARK_GRAY),
           align=_align("left", "center", indent=1))
    _row_h(ws, 4, 10)

    current_row = 5

    adjusted_result = None
    if cox_results and "adjusted" in cox_results:
        candidate = cox_results["adjusted"]
        if candidate and "error" not in candidate and "treatment_hr" in candidate:
            adjusted_result = candidate

    if adjusted_result is not None:
        covariates = adjusted_result.get("covariates") or COVARIATES_ADJUSTED
        tx_hr = adjusted_result["treatment_hr"]
        n_patients = adjusted_result.get("n_patients", "—")

        _row_h(ws, current_row, 18)
        _cell(ws, current_row, 2,
              f"Adjusted for {_format_adjusted_covariates(covariates)}",
              font=_font(size=10, bold=True, color=C_DARK_GRAY),
              align=_align("left", "center", indent=1))
        _cell(ws, current_row, 3, f"n = {n_patients}",
              font=_font(size=10, bold=True, color=C_DARK_GRAY),
              align=_align("left", "center"))
        current_row += 1

        _row_h(ws, current_row, 22)
        featured_headers = ["Treatment Comparison", "HR (95% CI)", "p-value"]
        for ci, col_name in enumerate(featured_headers):
            _cell(ws, current_row, 2 + ci, col_name,
                  font=_font(size=10, bold=True, color=C_WHITE),
                  fill=_fill(C_BLUE),
                  align=_align("center", "center"),
                  border=_thin_border())
        current_row += 1

        hr_ci = (
            f"{tx_hr['HR']:.3f} ({tx_hr['HR_lower']:.3f}, {tx_hr['HR_upper']:.3f})"
        )
        p_val = tx_hr["p_value"]
        p_str = "<0.001" if p_val < 0.001 else f"{p_val:.4f}"

        _row_h(ws, current_row, 20)
        featured_fill = _fill(C_LIGHT_BLUE)
        _cell(ws, current_row, 2, "Continuation vs Fixed-Duration",
              font=_font(size=10, bold=True, color=C_NAVY),
              fill=featured_fill,
              align=_align("left", "center", indent=1),
              border=_thin_border())
        _cell(ws, current_row, 3, hr_ci,
              font=_font(size=10, bold=True, color=C_NAVY),
              fill=featured_fill,
              align=_align("center", "center"),
              border=_thin_border())
        _cell(ws, current_row, 4, p_str,
              font=_font(size=10, bold=True, color=C_NAVY),
              fill=featured_fill,
              align=_align("center", "center"),
              border=_thin_border())
        current_row += 2

    if comparison_df is None or comparison_df.empty:
        _cell(ws, current_row, 2, "No Cox models available.",
              font=_font(size=10, italic=True, color=C_DARK_GRAY))
        return

    # Column headers
    cols = list(comparison_df.columns)
    _row_h(ws, current_row, 22)
    for ci, col_name in enumerate(cols):
        _cell(ws, current_row, 2 + ci, col_name,
              font=_font(size=10, bold=True, color=C_WHITE),
              fill=_fill(C_NAVY),
              align=_align("center", "center"),
              border=_thin_border())

    # Data rows
    for ri, (_, row_data) in enumerate(comparison_df.iterrows()):
        r = current_row + 1 + ri
        _row_h(ws, r, 18)
        fill = _fill(C_OFF_WHITE) if ri % 2 == 0 else _fill(C_WHITE)
        for ci, col_name in enumerate(cols):
            val = row_data[col_name]
            val_str = str(val) if val is not None else ""
            is_model = ci == 0
            _cell(ws, r, 2 + ci, val_str,
                  font=_font(size=10, bold=is_model),
                  fill=fill,
                  align=_align("left" if is_model else "center", "center"),
                  border=_thin_border(bottom_only=True))

    # Embed forest plot image
    img_path = Path(OUTPUT_DIR) / "forest_model_comparison.png"
    img_row = current_row + 1 + len(comparison_df) + 2
    if img_path.exists():
        try:
            img = XLImage(str(img_path))
            img.width = 640
            img.height = 340
            ws.add_image(img, f"B{img_row}")
        except Exception:
            pass

    ws.freeze_panes = f"B{current_row + 1}"


def _build_subgroup_sheet(wb, subgroup_df):
    """Sheet: Subgroup Analysis."""
    ws = wb.create_sheet("Subgroup Analysis")
    ws.sheet_properties.tabColor = "4393C3"
    _no_gridlines(ws)

    _col(ws, 1, 3)
    _col(ws, 2, 26)
    _col(ws, 3, 26)
    for c in range(4, 11):
        _col(ws, c, 16)
    _col(ws, 11, 3)

    _row_h(ws, 1, 8)
    _row_h(ws, 2, 30)
    _merge(ws, 2, 2, 10, "Subgroup Analysis — Treatment HR by Patient Characteristics",
           font=_font(size=14, bold=True, color=C_WHITE),
           fill=_fill(C_NAVY),
           align=_align("left", "center", indent=1))
    _row_h(ws, 3, 16)
    _merge(ws, 3, 2, 10,
           "Unadjusted Cox within each subgroup. Interaction p-value tests heterogeneity.",
           font=_font(size=10, italic=True, color=C_DARK_GRAY),
           align=_align("left", "center", indent=1))
    _row_h(ws, 4, 10)

    if subgroup_df is None or subgroup_df.empty:
        _cell(ws, 5, 2, "No subgroup results available.",
              font=_font(size=10, italic=True, color=C_DARK_GRAY))
        return

    display_cols = ["Variable", "Level", "N", "Events", "HR", "HR_lower",
                    "HR_upper", "p_value", "interaction_p"]
    cols = [c for c in display_cols if c in subgroup_df.columns]

    _row_h(ws, 5, 22)
    headers_display = {
        "Variable": "Variable", "Level": "Level", "N": "N",
        "Events": "Events", "HR": "HR", "HR_lower": "95% CI Lower",
        "HR_upper": "95% CI Upper", "p_value": "p-value",
        "interaction_p": "Interaction p",
    }
    for ci, col_name in enumerate(cols):
        _cell(ws, 5, 2 + ci, headers_display.get(col_name, col_name),
              font=_font(size=10, bold=True, color=C_WHITE),
              fill=_fill(C_NAVY),
              align=_align("center", "center"),
              border=_thin_border())

    for ri, (_, row_data) in enumerate(subgroup_df.iterrows()):
        r = 6 + ri
        _row_h(ws, r, 18)
        is_overall = row_data.get("Variable") == "Overall"
        fill = _fill(C_LIGHT_BLUE) if is_overall else (
            _fill(C_OFF_WHITE) if ri % 2 == 0 else _fill(C_WHITE))

        for ci, col_name in enumerate(cols):
            val = row_data.get(col_name, "")
            if isinstance(val, float) and not pd.isna(val):
                val_str = f"{val:.3f}" if col_name in ["HR", "HR_lower", "HR_upper",
                                                        "p_value", "interaction_p"] else f"{val:.0f}"
            elif pd.isna(val):
                val_str = ""
            else:
                val_str = str(val)

            _cell(ws, r, 2 + ci, val_str,
                  font=_font(size=10, bold=is_overall),
                  fill=fill,
                  align=_align("left" if ci < 2 else "center", "center"),
                  border=_thin_border(bottom_only=True))

    # Embed forest plot
    img_path = Path(OUTPUT_DIR) / "forest_subgroup_analysis.png"
    img_row = 6 + len(subgroup_df) + 2
    if img_path.exists():
        try:
            img = XLImage(str(img_path))
            img.width = 680
            img.height = 400
            ws.add_image(img, f"B{img_row}")
        except Exception:
            pass

    ws.freeze_panes = "B6"


def _build_full_cox_sheet(wb, cox_results):
    """Sheet: Cox Primary Model — full covariate table + forest plot."""
    ws = wb.create_sheet("Cox Primary Model")
    ws.sheet_properties.tabColor = "B2182B"
    _no_gridlines(ws)

    _col(ws, 1, 3)
    _col(ws, 2, 36)   # Covariate
    _col(ws, 3, 14)   # HR
    _col(ws, 4, 16)   # 95% CI
    _col(ws, 5, 12)   # p-value
    _col(ws, 6, 3)

    _row_h(ws, 1, 8)
    _row_h(ws, 2, 30)
    _merge(ws, 2, 2, 5, "Cox Proportional Hazards — Primary Model (All Covariates)",
           font=_font(size=14, bold=True, color=C_WHITE),
           fill=_fill(C_NAVY),
           align=_align("left", "center", indent=1))
    _row_h(ws, 3, 16)
    _merge(ws, 3, 2, 5,
           "Fully adjusted Cox model. Reference: Continuation vs Fixed-Duration. HR <1 favours Continuation.",
           font=_font(size=10, italic=True, color=C_DARK_GRAY),
           align=_align("left", "center", indent=1))
    _row_h(ws, 4, 10)

    # Pick primary model: prefer "full", fallback to "clinical"
    primary_result = None
    primary_name = None
    for key in ["full", "clinical", "minimal", "unadjusted"]:
        if cox_results and key in cox_results and "error" not in cox_results[key]:
            primary_result = cox_results[key]
            primary_name = key.title()
            break

    if primary_result is None:
        _merge(ws, 5, 2, 5, "No Cox model results available.",
               font=_font(size=10, italic=True, color=C_DARK_GRAY),
               align=_align("left", "center", indent=1))
        return

    summary_df = primary_result["summary"].copy()

    # Annotation row
    _row_h(ws, 4, 16)
    concordance = primary_result.get("concordance", float("nan"))
    n_pts = primary_result.get("n_patients", "?")
    n_evt = primary_result.get("n_events", "?")
    _merge(ws, 4, 2, 5,
           f"Model: {primary_name}  |  N={n_pts}  |  Events={n_evt}  |  C-index={concordance:.3f}",
           font=_font(size=9, bold=True, color=C_DARK_GRAY),
           fill=_fill(C_LIGHT_GRAY),
           align=_align("left", "center", indent=1))

    # Column headers
    _row_h(ws, 5, 22)
    headers = [("Covariate", "left"), ("HR", "center"),
               ("95% CI", "center"), ("p-value", "center")]
    for ci, (hdr, aln) in enumerate(headers):
        _cell(ws, 5, 2 + ci, hdr,
              font=_font(size=10, bold=True, color=C_WHITE),
              fill=_fill(C_NAVY),
              align=_align(aln, "center", indent=1 if aln == "left" else 0),
              border=_thin_border())

    # Treatment row first (highlight it)
    tx_row_idx = None
    if "treatment_continuation" in summary_df.index:
        tx_row_idx = list(summary_df.index).index("treatment_continuation")

    for ri, (idx, row_data) in enumerate(summary_df.iterrows()):
        r = 6 + ri
        _row_h(ws, r, 18)
        is_treatment = (idx == "treatment_continuation")

        hr_val   = row_data.get("exp(coef)", float("nan"))
        hr_lo    = row_data.get("exp(coef) lower 95%", float("nan"))
        hr_hi    = row_data.get("exp(coef) upper 95%", float("nan"))
        p_val    = row_data.get("p", float("nan"))

        hr_str = f"{hr_val:.3f}" if not np.isnan(hr_val) else "—"
        ci_str = (f"{hr_lo:.3f} – {hr_hi:.3f}"
                  if not (np.isnan(hr_lo) or np.isnan(hr_hi)) else "—")
        p_str  = (f"{p_val:.4f}" if not np.isnan(p_val) else "—")

        label = "Continuation vs Fixed-Duration (Treatment)" if is_treatment else str(idx)

        if is_treatment:
            fill = _fill(C_LIGHT_BLUE)
            fnt_lbl = _font(size=10, bold=True, color=C_NAVY)
            fnt_val = _font(size=10, bold=True, color=C_NAVY)
        elif ri % 2 == 0:
            fill = _fill(C_OFF_WHITE)
            fnt_lbl = fnt_val = _font(size=10)
        else:
            fill = _fill(C_WHITE)
            fnt_lbl = fnt_val = _font(size=10)

        _cell(ws, r, 2, label, font=fnt_lbl, fill=fill,
              align=_align("left", "center", indent=1),
              border=_thin_border(bottom_only=True))
        _cell(ws, r, 3, hr_str, font=fnt_val, fill=fill,
              align=_align("center", "center"),
              border=_thin_border(bottom_only=True))
        _cell(ws, r, 4, ci_str, font=fnt_val, fill=fill,
              align=_align("center", "center"),
              border=_thin_border(bottom_only=True))

        # Highlight significant p-values
        p_color = C_CONT if (not np.isnan(p_val) and p_val < 0.05) else "000000"
        _cell(ws, r, 5, p_str,
              font=_font(size=10, bold=(not np.isnan(p_val) and p_val < 0.05),
                         color=p_color),
              fill=fill,
              align=_align("center", "center"),
              border=_thin_border(bottom_only=True))

    # Embed forest plot if it exists
    img_row = 6 + len(summary_df) + 2
    for fname in [f"forest_plot_full_cox.png", "forest_plot_clinical_cox.png",
                  "forest_model_comparison.png"]:
        img_path = Path(OUTPUT_DIR) / fname
        if img_path.exists():
            try:
                img = XLImage(str(img_path))
                img.width  = 640
                img.height = 400
                ws.add_image(img, f"B{img_row}")
            except Exception:
                pass
            break

    ws.freeze_panes = "B6"


def _build_landmark_sheet(wb, landmark_df):
    """Sheet: Landmark Sensitivity Analysis."""
    ws = wb.create_sheet("Landmark Sensitivity")
    ws.sheet_properties.tabColor = C_GOLD
    _no_gridlines(ws)

    _col(ws, 1, 3)
    for c in range(2, 12):
        _col(ws, c, 18)
    _col(ws, 12, 3)

    _row_h(ws, 1, 8)
    _row_h(ws, 2, 30)
    _merge(ws, 2, 2, 11, "Landmark Sensitivity Analysis",
           font=_font(size=14, bold=True, color=C_WHITE),
           fill=_fill(C_NAVY),
           align=_align("left", "center", indent=1))
    _row_h(ws, 3, 16)
    _merge(ws, 3, 2, 11,
           "Treatment HR re-estimated at different landmark timepoints (27, 29, 32 months).",
           font=_font(size=10, italic=True, color=C_DARK_GRAY),
           align=_align("left", "center", indent=1))
    _row_h(ws, 4, 10)

    if landmark_df is None or landmark_df.empty:
        _cell(ws, 5, 2, "No landmark sensitivity results available.",
              font=_font(size=10, italic=True, color=C_DARK_GRAY))
        return

    cols = list(landmark_df.columns)
    _row_h(ws, 5, 22)
    for ci, col_name in enumerate(cols):
        _cell(ws, 5, 2 + ci, col_name,
              font=_font(size=10, bold=True, color=C_WHITE),
              fill=_fill(C_NAVY),
              align=_align("center", "center"),
              border=_thin_border())

    for ri, (_, row_data) in enumerate(landmark_df.iterrows()):
        r = 6 + ri
        _row_h(ws, r, 18)
        fill = _fill(C_OFF_WHITE) if ri % 2 == 0 else _fill(C_WHITE)
        for ci, col_name in enumerate(cols):
            val = row_data[col_name]
            if isinstance(val, float):
                val_str = f"{val:.3f}" if col_name in ["HR", "HR_lower", "HR_upper",
                                                        "p_value", "C-index"] else f"{val:.0f}"
            else:
                val_str = str(val)
            _cell(ws, r, 2 + ci, val_str,
                  font=_font(size=10),
                  fill=fill,
                  align=_align("center", "center"),
                  border=_thin_border(bottom_only=True))

    # Embed plot
    img_path = Path(OUTPUT_DIR) / "landmark_sensitivity.png"
    img_row = 6 + len(landmark_df) + 2
    if img_path.exists():
        try:
            img = XLImage(str(img_path))
            img.width = 560
            img.height = 360
            ws.add_image(img, f"B{img_row}")
        except Exception:
            pass


def create_excel_report(cohort_df, attrition, km_output, km_supporting, table1_df,
                        cox_results=None, lasso_result=None, comparison_df=None,
                        subgroup_df=None, ph_result=None, landmark_df=None):
    """
    Build and save the full Excel presentation workbook.

    Parameters
    ----------
    cohort_df      : analysis cohort DataFrame
    attrition      : dict of patient counts at each step
    km_output      : dict from run_kaplan_meier()
    km_supporting  : dict from build_km_supporting_table()
    table1_df      : DataFrame from generate_table1()
    cox_results    : dict from run_multiple_cox_models() (optional)
    lasso_result   : dict from run_lasso_cox() (optional)
    comparison_df  : DataFrame from build_model_comparison_table() (optional)
    subgroup_df    : DataFrame from run_subgroup_analyses() (optional)
    ph_result      : dict from test_proportional_hazards() (optional)
    landmark_df    : DataFrame from run_landmark_sensitivity() (optional)
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Sheet order: Overview → Patient Flow → Table 1 → OS → Cox Primary → Cox Comparison → Subgroup → Landmark → Methodology
    print("[excel_report] Building Overview sheet...")
    _build_overview(wb, cohort_df, attrition, km_output)

    print("[excel_report] Building Patient Flow sheet...")
    _build_patient_flow(wb, attrition)

    print("[excel_report] Building Baseline Characteristics sheet...")
    _build_baseline(wb, table1_df)

    print("[excel_report] Building Overall Survival sheet...")
    _build_os_analysis(wb, cohort_df, km_output, km_supporting)

    # Cox primary model (all covariates) — always build if any cox results exist
    if cox_results:
        print("[excel_report] Building Cox Primary Model sheet...")
        _build_full_cox_sheet(wb, cox_results)

    # Cox model comparison (treatment HR across adjustment levels)
    if comparison_df is not None:
        print("[excel_report] Building Cox Model Comparison sheet...")
        _build_cox_comparison(wb, comparison_df, cox_results=cox_results)

    if subgroup_df is not None:
        print("[excel_report] Building Subgroup Analysis sheet...")
        _build_subgroup_sheet(wb, subgroup_df)

    if landmark_df is not None:
        print("[excel_report] Building Landmark Sensitivity sheet...")
        _build_landmark_sheet(wb, landmark_df)

    print("[excel_report] Building Methodology sheet...")
    _build_methodology(wb, cohort_df, attrition, km_output)

    out_path = Path(OUTPUT_DIR) / "pembro_nsclc_os_analysis.xlsx"
    wb.save(str(out_path))
    print(f"[excel_report] Workbook saved to {out_path}")
    return out_path
